# -*- coding: utf-8 -*-
"""
智慧招采平台功能清单生成器
覆盖全平台 14 个子系统、约 705 个功能点
生成文件：智慧招采平台功能清单.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# ============================================================
# 子系统编号体系（与 features_data/*.py 中的 SUBSYSTEM 一一对应）
# ============================================================
SUBSYSTEM_CODES = {
    "门户网站": "MENHU",
    "采购子系统": "PCG",
    "供应商子系统": "SPR",
    "企业管理后台": "ENT",
    "开标系统": "KAI",
    "评标系统": "PING",
    "运营系统": "OPS",
    "控制中心": "CTRL",
    "专家系统": "EXP",
    "编制工具系统": "TOOL",
    "大数据系统": "DATA",
    "数字证书管理后台": "CA",
    "统一接口平台": "API",
    "监督系统": "SUP",
}

# 子系统数据加载顺序
SUBSYSTEM_MODULES_ORDER = [
    "features_data.menhu",
    "features_data.pcg",
    "features_data.spr",
    "features_data.ent",
    "features_data.kai",
    "features_data.ping",
    "features_data.ops",
    "features_data.ctrl",
    "features_data.exp",
    "features_data.tool",
    "features_data.data_",
    "features_data.ca",
    "features_data.api",
    "features_data.sup",
]


# ============================================================
# 样式定义
# ============================================================
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="微软雅黑", size=10)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="top", wrap_text=True)

ROW_FILL_P0 = PatternFill("solid", fgColor="FDE0D8")  # 淡红
ROW_FILL_P1 = PatternFill("solid", fgColor="FFF4CC")  # 淡黄
ROW_FILL_P2 = PatternFill("solid", fgColor="F2F2F2")  # 淡灰

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

COLUMN_WIDTHS = {
    "A": 14,  # 功能编号
    "B": 14,  # 子系统
    "C": 14,  # 模块
    "D": 16,  # 子模块
    "E": 22,  # 功能名称
    "F": 50,  # 功能描述
    "G": 30,  # 业务规则
    "H": 18,  # 操作角色
    "I": 8,  # 优先级
    "J": 10,  # 开发状态
    "K": 10,  # 需求章节
}

HEADERS = [
    "功能编号",
    "子系统",
    "模块",
    "子模块",
    "功能名称",
    "功能描述",
    "业务规则/强控",
    "操作角色",
    "优先级",
    "开发状态",
    "需求章节",
]


# ============================================================
# Excel 生成逻辑
# 数据元组结构：(模块, 子模块, 功能名称, 功能描述, 业务规则, 操作角色, 优先级, 章节)
# ============================================================
def import_features(module_name):
    import importlib

    mod = importlib.import_module(module_name)
    return mod.SUBSYSTEM, mod.FEATURES


def build_excel(output_path):
    import importlib.util

    base_dir = Path(__file__).parent / "features_data"
    for f in list(base_dir.glob("*.pyc")) + list(
        (base_dir / "__pycache__").glob("*.pyc")
        if (base_dir / "__pycache__").exists()
        else []
    ):
        try:
            f.unlink()
        except:
            pass

    all_rows = []
    for module_name in SUBSYSTEM_MODULES_ORDER:
        subsystem, features = import_features(module_name)
        code = SUBSYSTEM_CODES.get(subsystem, subsystem)

        # 按模块+子模块分组编号
        groups = {}
        for item in features:
            key = (item[0], item[1])  # (模块, 子模块)
            groups.setdefault(key, []).append(item)

        for i, (key, items) in enumerate(groups.items(), start=1):
            for j, item in enumerate(items, start=1):
                feature_code = f"{code}-{i:02d}.{j:02d}"
                all_rows.append(
                    (
                        feature_code,
                        subsystem,
                        item[0],
                        item[1],
                        item[2],
                        item[3],
                        item[4],
                        item[5],
                        item[6],
                        "未开始",
                        item[7],
                    )
                )

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "功能清单"

    # 列宽
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 写表头
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28

    # 写数据行
    for row_idx, row_data in enumerate(all_rows, start=2):
        priority = row_data[8]
        if priority == "P0":
            row_fill = ROW_FILL_P0
        elif priority == "P1":
            row_fill = ROW_FILL_P1
        else:
            row_fill = ROW_FILL_P2

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = (
                CENTER_ALIGNMENT if col_idx in (1, 9, 10) else BODY_ALIGNMENT
            )

    # 冻结表头 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 保存
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # 统计输出
    print(f"✅ 生成成功: {output_path}")
    print(f"   共 {len(all_rows)} 条功能记录，覆盖 {len(SUBSYSTEM_CODES)} 个子系统")

    stats = {}
    for r in all_rows:
        stats[r[1]] = stats.get(r[1], 0) + 1
    print("\n子系统功能分布:")
    for subsys, cnt in stats.items():
        code = SUBSYSTEM_CODES.get(subsys, subsys)
        print(f"  [{code:<6}] {subsys:<10}  {cnt} 个功能点")


if __name__ == "__main__":
    output = Path(__file__).parent / "智慧招采平台功能清单.xlsx"
    build_excel(output)
